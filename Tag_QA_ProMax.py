import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import csv
import asyncio
from playwright.async_api import async_playwright
import threading
import os
import shutil
import json
import tempfile
from datetime import datetime
from collections import deque
import time
import sys
from pathlib import Path
from PIL import Image, ImageGrab, ImageDraw, ImageFont
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage

class TaggingAutomationApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Tagging Automation QA Pro")
        self.root.geometry("1200x700")

                # --- Workspace Path Setup ---
        self._setup_workspace_paths(Path.home() / "Documents")

        # --- Constants & State ---
        self.LANG_MAP = {
            "Traditional Chinese": "tc",
            "Simplified Chinese": "sc",
            "English": "en"
        }
        self.LANG_MAP_INV = {v: k for k, v in self.LANG_MAP.items()}
        self.is_updating_ui = False

        # --- Style Definitions ---
        style = ttk.Style()
        style.configure("Highlight.TButton", font=("Helvetica", 12, "bold"))

        # Instance variables
        self.playwright_loop = None
        self.browser_context = None
        self.playwright_page = None
        self.all_logs = []
        self.keyword_matches = {}
        self.update_timer = None
        self.active_filter_keyword = None
        self.last_clicked_keyword_index = None
        self.urls = [{'url': 'https://www.google.com', 'lang': 'en', 'num': 1}] # Now a list of objects
        self.report_data = []
        
        # Undo/Redo stacks
        self.undo_stack = deque(maxlen=5)
        self.redo_stack = deque()

        # --- Main Layout Frames ---
        top_controls_frame = ttk.Frame(root, padding="10")
        top_controls_frame.pack(fill=tk.X)

        main_paned_window = ttk.PanedWindow(root, orient=tk.HORIZONTAL)
        main_paned_window.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))

        left_pane = ttk.Frame(main_paned_window, padding=5)
        main_paned_window.add(left_pane, weight=1)

        right_pane = ttk.Frame(main_paned_window, padding=5)
        main_paned_window.add(right_pane, weight=4) # Give more space to the log

        # --- Populate Top Controls ---
        self.setup_top_controls(top_controls_frame)

        # --- Populate Left Pane (Keyword Management) ---
        self.setup_keyword_pane(left_pane)

        # --- Populate Right Pane (Unified Log View) ---
        self.setup_log_pane(right_pane)

        # --- Status Bar ---
        self.status_var = tk.StringVar()
        self.status_label = ttk.Label(root, textvariable=self.status_var, padding=5, relief=tk.SUNKEN)
        self.status_label.pack(side=tk.BOTTOM, fill=tk.X)
        self.update_status("Ready")

        # Initialize URL combobox with correctly formatted strings
        self.update_urls(self.urls)

    def setup_top_controls(self, parent_frame):
        # URL Frame
        url_frame = ttk.Frame(parent_frame)
        url_frame.pack(fill=tk.X, expand=True, pady=(0, 5))
        ttk.Label(url_frame, text="URL:").pack(side=tk.LEFT)
        self.url_var = tk.StringVar()
        self.url_combobox = ttk.Combobox(url_frame, textvariable=self.url_var, values=self.urls)
        self.url_combobox.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        self.url_combobox.set(self.urls[0] if self.urls else "")
        manage_urls_button = ttk.Button(url_frame, text="Manage URLs", command=self.open_url_manager)
        manage_urls_button.pack(side=tk.LEFT)
        change_workspace_button = ttk.Button(url_frame, text="Change Workspace", command=self.change_workspace)
        change_workspace_button.pack(side=tk.LEFT, padx=5)

        # Browser Control Frame
        browser_control_frame = ttk.Frame(parent_frame)
        browser_control_frame.pack(fill=tk.X, expand=True)
        self.mode_var = tk.StringVar(value="Incognito")
        mode_menu = ttk.OptionMenu(browser_control_frame, self.mode_var, "Incognito", "Normal", "Incognito")
        mode_menu.pack(side=tk.LEFT)
        self.browser_button = ttk.Button(browser_control_frame, text="Start Browser", command=self.toggle_browser)
        self.browser_button.pack(side=tk.LEFT, padx=5)
        self.test_button = ttk.Button(browser_control_frame, text="Test", command=self.start_test_thread)
        self.test_button.pack(side=tk.LEFT, padx=5)
        self.export_button = ttk.Button(browser_control_frame, text="Export Log", command=self.export_logs)
        self.export_button.pack(side=tk.LEFT, padx=5)
        self.clear_button = ttk.Button(browser_control_frame, text="Clear All", command=self.clear_all)
        self.clear_button.pack(side=tk.LEFT, padx=5)
        self.screenshot_button = ttk.Button(browser_control_frame, text="Screenshot", command=self.capture_and_stitch_thread)
        self.screenshot_button.pack(side=tk.LEFT, padx=5)
        self.output_button = ttk.Button(browser_control_frame, text="Output", command=self.generate_excel_report)
        self.output_button.pack(side=tk.LEFT, padx=5)
        self.fast_test_button = ttk.Button(
            browser_control_frame,
            text="Fast Test!",
            command=self.start_fast_test_thread,
            style="Highlight.TButton"
        )
        self.fast_test_button.pack(side=tk.LEFT, padx=(30, 5))


    def _setup_workspace_paths(self, parent_dir):
        """Initializes or updates all workspace-related paths."""
        self.workspace_parent_dir = parent_dir
        self.base_dir = self.workspace_parent_dir / "Tag_QA_Files"
        self.captures_dir = self.base_dir / "Pictures"
        self.sessions_dir = self.base_dir / "Sessions"
        self.logs_dir = self.base_dir / "Logs"
        self.outputs_dir = self.base_dir / "Outputs"

        # Create directories if they don't exist
        self.captures_dir.mkdir(parents=True, exist_ok=True)
        self.sessions_dir.mkdir(parents=True, exist_ok=True)
        self.logs_dir.mkdir(parents=True, exist_ok=True)
        self.outputs_dir.mkdir(parents=True, exist_ok=True)

    def change_workspace(self):
        """Opens a dialog to move the workspace to a new directory."""
        new_parent_dir = filedialog.askdirectory(
            title="Select New Parent Directory for Tag_QA_Files",
            initialdir=self.workspace_parent_dir
        )

        if not new_parent_dir:
            return # User cancelled

        new_parent_path = Path(new_parent_dir)
        if new_parent_path == self.workspace_parent_dir:
            messagebox.showinfo("Info", "This is already the current workspace directory.")
            return

        target_path = new_parent_path / "Tag_QA_Files"
        if target_path.exists():
            messagebox.showerror("Error", f"The destination directory already contains a 'Tag_QA_Files' folder. Please choose a different location or remove the existing one.")
            return

        try:
            self.update_status(f"Moving workspace to {new_parent_path}...")
            shutil.move(str(self.base_dir), str(new_parent_path))
            
            # Update all internal paths to point to the new location
            self._setup_workspace_paths(new_parent_path)
            
            self.update_status("Workspace moved successfully.")
            messagebox.showinfo("Success", f"Workspace has been moved to:\n{self.base_dir}")
        except Exception as e:
            self.update_status(f"Error moving workspace: {e}")
            messagebox.showerror("Move Error", f"Failed to move workspace: {e}")


    def open_url_manager(self):
        URLManager(self.root, self, self.urls, self.update_urls)

    def update_urls(self, new_urls):
        self.urls = new_urls
        display_urls = [f"[{u['num']}] [{u['lang']}] {u['url']}" for u in self.urls]
        self.url_combobox['values'] = display_urls
        if display_urls:
            self.url_combobox.set(display_urls[0])
        else:
            self.url_combobox.set("")

    def toggle_keep_on_top(self):
        """Toggles the window's always-on-top attribute."""
        is_on_top = self.keep_on_top_var.get()
        self.root.attributes("-topmost", is_on_top)

    def setup_keyword_pane(self, parent_frame):
        self.keyword_text_var = tk.StringVar()
        # Frame for keyword input
        keyword_input_frame = ttk.Frame(parent_frame)
        keyword_input_frame.pack(fill=tk.X, pady=(0, 5))

        # Top row for keyword entry
        top_row_frame = ttk.Frame(keyword_input_frame)
        top_row_frame.pack(fill=tk.X, pady=(0, 5))

        ttk.Label(top_row_frame, text="Keyword:").pack(side=tk.LEFT, padx=(0, 5))
        self.keyword_entry = ttk.Entry(top_row_frame, textvariable=self.keyword_text_var)
        self.keyword_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)

        # Second row for ID, Num, and Language
        second_row_frame = ttk.Frame(keyword_input_frame)
        second_row_frame.pack(fill=tk.X)

        ttk.Label(second_row_frame, text="ID:").pack(side=tk.LEFT, padx=(0,5))
        self.button_id_var = tk.StringVar()
        self.button_id_entry = ttk.Entry(second_row_frame, textvariable=self.button_id_var, width=15)
        self.button_id_entry.pack(side=tk.LEFT, padx=(0,5))

        self.keyword_num_var = tk.StringVar(value="1")
        ttk.Label(second_row_frame, text="Num:").pack(side=tk.LEFT, padx=(5,0))
        self.keyword_num_entry = ttk.Entry(second_row_frame, textvariable=self.keyword_num_var, width=5)
        self.keyword_num_entry.pack(side=tk.LEFT, padx=(0,5))

        self.lang_var = tk.StringVar(value="Traditional Chinese")
        lang_options = list(self.LANG_MAP.keys())
        self.lang_combobox = ttk.Combobox(second_row_frame, textvariable=self.lang_var, values=lang_options, state="readonly", width=20)
        self.lang_combobox.pack(side=tk.RIGHT, padx=5)
        
        # Bottom row for the Add button
        add_button = ttk.Button(keyword_input_frame, text="Add", command=self.add_keyword)
        add_button.pack(fill=tk.X, expand=True, pady=(5, 0))


        # Listbox for keywords
        listbox_frame = ttk.Frame(parent_frame)
        listbox_frame.pack(fill=tk.BOTH, expand=True)
        self.keyword_listbox = tk.Listbox(listbox_frame, selectmode=tk.BROWSE)
        self.keyword_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        list_scrollbar = ttk.Scrollbar(listbox_frame, orient=tk.VERTICAL, command=self.keyword_listbox.yview)
        list_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.keyword_listbox.config(yscrollcommand=list_scrollbar.set)

        # Bind events
        self.keyword_listbox.bind("<ButtonRelease-1>", self._on_keyword_click)
        self.keyword_listbox.bind("<Control-v>", self.handle_paste_event)
        self.keyword_listbox.bind("<Command-v>", self.handle_paste_event)
        self.keyword_listbox.bind("<Delete>", self.remove_selected_keyword_event)
        self.keyword_listbox.bind("<Command-BackSpace>", self.remove_selected_keyword_event)

        # Bind auto-update traces
        self.keyword_text_var.trace_add("write", self.handle_keyword_update)
        self.lang_var.trace_add("write", self.handle_keyword_update)
        self.keyword_num_var.trace_add("write", self.handle_keyword_update)
        self.button_id_var.trace_add("write", self.handle_keyword_update)

        # Undo/Redo bindings
        self.root.bind("<Control-z>", self.undo_keywords)
        self.root.bind("<Command-z>", self.undo_keywords)
        self.root.bind("<Control-y>", self.redo_keywords)
        self.root.bind("<Command-y>", self.redo_keywords)

        remove_button = ttk.Button(parent_frame, text="Remove All", command=self.remove_all_keywords)
        remove_button.pack(fill=tk.X, pady=5)

        # Save/Load buttons
        session_button_frame = ttk.Frame(parent_frame)
        session_button_frame.pack(fill=tk.X, pady=(5,0))
        save_button = ttk.Button(session_button_frame, text="Save Session", command=self.save_session)
        save_button.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=(0,2))
        load_button = ttk.Button(session_button_frame, text="Load Session", command=self.load_session)
        load_button.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=(2,0))

        # --- Help/Shortcuts Section ---
        self.setup_shortcuts_pane(parent_frame)

        # Save initial state for undo
        self._save_keyword_state()

    def setup_shortcuts_pane(self, parent_frame):
        """Creates a collapsible pane for shortcuts and help text."""
        help_frame = ttk.Frame(parent_frame, padding=10, relief="groove", borderwidth=1)
        help_frame.pack(fill=tk.X, pady=(10, 0))

        shortcuts_text = """
        Shortcuts
        Paste:  Cmd/Ctrl+V
        Delete: Delete/Backspace
        Undo:   Cmd/Ctrl+Z
        Redo:   Cmd/Ctrl+Y
        """

        help_label = ttk.Label(help_frame, text=shortcuts_text.strip(), justify=tk.LEFT)
        help_label.pack(fill=tk.X)

    def setup_log_pane(self, parent_frame):
        columns = ("name", "status", "method", "type", "size", "time", "url_hash")
        self.log_tree = ttk.Treeview(parent_frame, columns=columns, show="headings")
        
        self.log_tree.heading("name", text="Name", command=lambda: self.sort_treeview("name", False))
        self.log_tree.heading("status", text="Status", command=lambda: self.sort_treeview("status", False))
        self.log_tree.heading("method", text="Method", command=lambda: self.sort_treeview("method", False))
        self.log_tree.heading("type", text="Type", command=lambda: self.sort_treeview("type", False))
        self.log_tree.heading("size", text="Size", command=lambda: self.sort_treeview("size", False))
        self.log_tree.heading("time", text="Time", command=lambda: self.sort_treeview("time", False))
        self.log_tree.heading("url_hash", text="URL Hash", command=lambda: self.sort_treeview("url_hash", False))

        for col in columns:
            self.log_tree.column(col, width=100, stretch=tk.YES)

        self.log_tree.column("name", width=100)
        self.log_tree.column("status", width=60)
        self.log_tree.column("method", width=60)
        self.log_tree.column("time", width=150)
        self.log_tree.column("url_hash", width=90)

        scrollbar = ttk.Scrollbar(parent_frame, orient=tk.VERTICAL, command=self.log_tree.yview)
        self.log_tree.configure(yscroll=scrollbar.set)
        
        self.log_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    def export_logs(self):
        if not self.all_logs:
            messagebox.showwarning("No Data", "There is no log data to export.")
            return

        default_filename = f"network_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        filepath = filedialog.asksaveasfilename(
            title="Save Log As",
            initialdir=self.logs_dir,
            initialfile=default_filename,
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )

        if not filepath:
            return # User cancelled

        try:
            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                # Write header
                writer.writerow([self.log_tree.heading(c)["text"] for c in self.log_tree["columns"]])
                # Write data
                writer.writerows(self.all_logs)
            
            messagebox.showinfo("Success", f"Log successfully exported to {os.path.basename(filepath)}")
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export log: {e}")

    def capture_and_stitch_thread(self):
        """Wrapper to run the async capture method from a sync button click."""
        if not self.playwright_loop or not self.playwright_loop.is_running():
            messagebox.showwarning("Not Ready", "The browser automation loop is not running.")
            return
        asyncio.run_coroutine_threadsafe(self.capture_and_stitch(), self.playwright_loop)

    async def capture_and_stitch(self, output_path=None, show_success_message=True):
        if not self.playwright_page or self.playwright_page.is_closed():
            self.root.after(0, lambda: messagebox.showwarning("Browser Not Ready", "Please start the browser first."))
            return

        loop = asyncio.get_running_loop()
        current_url = self.playwright_page.url

        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            temp_dir = tempfile.gettempdir()
            browser_shot_path = os.path.join(temp_dir, f"temp_browser_{timestamp}.png")
            gui_shot_path = os.path.join(temp_dir, f"temp_gui_{timestamp}.png")
            
            if output_path is None:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                output_path = self.captures_dir / f"stitched_capture_{timestamp}.png"

            # --- Capture Browser --- #
            await self.playwright_page.bring_to_front()
            await self.playwright_page.evaluate("window.scrollTo(0, 0)")
            await asyncio.sleep(0.3) # Wait for focus and scroll
            await self.playwright_page.screenshot(path=browser_shot_path)

            # --- Capture GUI --- #
            def grab_gui():
                self.root.attributes("-topmost", True)
                self.root.update_idletasks()
                time.sleep(0.3) # Wait for window to come to front
                x, y, width, height = self.root.winfo_rootx(), self.root.winfo_rooty(), self.root.winfo_width(), self.root.winfo_height()
                ImageGrab.grab(bbox=(x, y, x + width, y + height)).save(gui_shot_path)
                self.root.attributes("-topmost", False)
            
            await loop.run_in_executor(None, grab_gui)

            # --- Stitch Images --- #
            def stitch_images_with_url(url_text):
                with Image.open(gui_shot_path) as gui_img, Image.open(browser_shot_path) as browser_img:
                    total_width = gui_img.width + browser_img.width
                    max_height = max(gui_img.height, browser_img.height)
                    
                    stitched_image = Image.new('RGB', (total_width, max_height), (255, 255, 255))
                    stitched_image.paste(gui_img, (0, 0))
                    stitched_image.paste(browser_img, (gui_img.width, 0))
                    stitched_image.save(output_path)

                os.remove(browser_shot_path)
                os.remove(gui_shot_path)
            
            await loop.run_in_executor(None, stitch_images_with_url, current_url)
            
            if show_success_message:
                self.root.after(0, lambda: messagebox.showinfo("Success", f"Screenshot saved to {os.path.basename(output_path)}"))
            else:
                self.update_status(f"Screenshot saved: {os.path.basename(output_path)}")

        except Exception as e:
            print(f"Capture Error: {e}")
            self.root.after(0, lambda: messagebox.showerror("Capture Error", f"An error occurred: {e}"))

    def generate_excel_report(self):
        """Generates and saves an Excel report from the collected report_data."""
        if not self.report_data:
            messagebox.showwarning("No Data", "No report data found. Please run the Fast Test first.")
            return

        self.update_status("Generating Excel report...")
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Test Report"

            # --- Sorting Data ---
            keyword_objects_order = self._get_keyword_objects()
            keyword_sort_map = {obj['text']: i for i, obj in enumerate(keyword_objects_order)}
            url_order = [u['url'] for u in self.urls]
            url_sort_map = {url: i for i, url in enumerate(url_order)}

            sorted_report_data = sorted(
                self.report_data,
                key=lambda item: (
                    keyword_sort_map.get(item['keyword'], 999),
                    url_sort_map.get(item['url'], 999)
                )
            )

            # --- Headers ---
            headers = ["Keyword", "Language", "Status", "URL", "Screenshot"]
            ws.append(headers)
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = cell.font.copy(bold=True)
            
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 10
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 40
            ws.column_dimensions['E'].width = 80 # Approx 600px

            # --- Populating Data ---
            for row_idx, item in enumerate(sorted_report_data, 2):
                ws.cell(row=row_idx, column=1, value=item['keyword'])
                ws.cell(row=row_idx, column=2, value=item['lang'])
                ws.cell(row=row_idx, column=3, value=item['status'])
                ws.cell(row=row_idx, column=4, value=item['url'])

                img_path = item['screenshot_path']
                if os.path.exists(img_path):
                    try:
                        img = OpenpyxlImage(img_path)
                        # Scale image to a fixed width, preserving aspect ratio
                        scale_width = 600
                        img.height = img.height * (scale_width / img.width)
                        img.width = scale_width
                        
                        ws.add_image(img, f'E{row_idx}')
                        ws.row_dimensions[row_idx].height = img.height * 0.75 # Convert pixels to points
                    except Exception as img_e:
                        ws.cell(row=row_idx, column=5, value=f"Error loading image: {img_e}")
                else:
                    ws.cell(row=row_idx, column=5, value="Image not found")

            # --- Save File ---
            report_filename = f"Test_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            report_path = self.outputs_dir / report_filename
            wb.save(report_path)
            self.update_status(f"Report saved: {report_path}")
            messagebox.showinfo("Success", f"Excel report saved as {report_path}")

        except Exception as e:
            self.update_status(f"Error generating report: {e}")
            messagebox.showerror("Report Error", f"Failed to generate Excel report: {e}")


    # --- Fast Test Automation --- #

    def update_status(self, message):
        """Thread-safe method to update the status bar."""
        self.root.after(0, lambda: self.status_var.set(message))

    def toggle_controls(self, enabled):
        """Enable or disable all major UI controls."""
        state = tk.NORMAL if enabled else tk.DISABLED
        # Toggle all buttons in the top control frame
        for child in self.browser_button.master.winfo_children():
            if isinstance(child, (ttk.Button, ttk.OptionMenu)):
                child.config(state=state)
        # Re-enable the fast test button specifically if it's the end
        self.fast_test_button.config(state=state)
        # Toggle keyword controls
        for child in self.keyword_entry.master.winfo_children():
            if isinstance(child, ttk.Button):
                child.config(state=state)
        self.keyword_listbox.config(state=state)
        self.keyword_entry.config(state="normal" if enabled else "disabled")

    def start_fast_test_thread(self):
        """Prepares UI and starts the automation process in a new thread."""
        if not self.urls:
            messagebox.showwarning("No URLs", "Please add at least one URL in Manage URLs.")
            return
        if not self._get_raw_keywords():
            messagebox.showwarning("No Keywords", "Please add at least one keyword.")
            return

        self.toggle_controls(False)
        self.update_status("Starting Fast Test...")
        self.report_data = [] # Clear previous report data
        
        thread = threading.Thread(target=self.run_full_automation, daemon=True)
        thread.start()

    def run_full_automation(self):
        """Sets up and runs the asyncio event loop for the automation."""
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        self.playwright_loop = loop
        try:
            loop.run_until_complete(self._orchestrate_all_urls())
            self.update_status("Fast Test Completed Successfully!")
        except Exception as e:
            print(f"Fast Test Error: {e}")
            self.update_status(f"Error: {e}")
            messagebox.showerror("Fast Test Error", f"An error occurred: {e}")
        finally:
            self.root.after(0, self.generate_excel_report)
            self.root.after(0, lambda: self.toggle_controls(True))
            self.root.after(0, lambda: self.browser_button.config(text="Start Browser")) # Reset button text
            loop.close()
            self.playwright_loop = None

    async def _orchestrate_all_urls(self):
        """Main async orchestrator to loop through URLs and run tests."""
        num_urls = len(self.urls)
        for i, url_obj in enumerate(self.urls):
            self.root.after(0, self.clear_all) # Clear logs and UI for the new run
            await asyncio.sleep(0.5) # Give a moment for UI to clear
            self.update_status(f"URL {i+1}/{num_urls}: Starting test for {url_obj['url']}")
            await self._automated_run_for_url(url_obj)

    async def _automated_run_for_url(self, url_obj):
        """Runs the full test-and-screenshot cycle for a single URL."""
        url_str = url_obj['url']
        url_lang = url_obj['lang']
        user_data_dir = tempfile.mkdtemp()
        clicked_button_ids_on_page = set()
        try:
            async with async_playwright() as p:
                # 1. Launch Browser
                self.update_status(f"Launching browser for {url_str}...")
                
                launch_options = {
                    "headless": False,
                    "args": ['--incognito'] if self.mode_var.get() == "Incognito" else [],
                }
                if sys.platform == "darwin": # macOS
                    chrome_path = "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"
                    if os.path.exists(chrome_path):
                        launch_options["executable_path"] = chrome_path

                context = await p.chromium.launch_persistent_context(
                    user_data_dir,
                    **launch_options
                )
                self.browser_context = context
                page = context.pages[0] if context.pages else await context.new_page()
                self.playwright_page = page
                page.on("response", self.handle_response)

                await page.goto(url_str, wait_until="domcontentloaded")

                # 2. Wait for initial page load to settle
                await self.wait_for_network_idle()

                # 3. Element test is now done per keyword if a button_id is present.

                # 5. Screenshot per relevant keyword
                all_keyword_objects = self._get_keyword_objects()
                relevant_keywords = [ 
                    kw for kw in all_keyword_objects 
                    if kw['lang'] == url_lang and kw.get('num', 1) == url_obj.get('num', 1)
                ]
                num_keywords = len(relevant_keywords)

                for i, keyword_obj in enumerate(relevant_keywords):
                    keyword_text = keyword_obj['text']
                    keyword_lang = keyword_obj['lang']
                    button_id = keyword_obj.get('button_id', '')

                    self.update_status(f"Processing keyword {i+1}/{num_keywords}: '{keyword_text}'...")

                    if button_id and button_id not in clicked_button_ids_on_page:
                        await self.click_button_by_id(button_id)
                        clicked_button_ids_on_page.add(button_id)
                        await self.wait_for_network_idle()

                    self.update_status(f"Capturing keyword {i+1}/{num_keywords}: '{keyword_text}' for URL lang '{url_lang}'...")
                    
                    select_keyword_event = threading.Event()
                    self.root.after(0, self._select_keyword_programmatically, keyword_text, select_keyword_event)
                    select_keyword_event.wait()
                    await asyncio.sleep(0.5)

                    sanitized_url = url_str.split('//')[-1].split('/')[0].replace('.', '_')
                    sanitized_keyword = keyword_text.replace(' ', '_').replace('/', '_')
                    filename = f"capture_{sanitized_url}_{sanitized_keyword}_{datetime.now().strftime('%Y%m%d%H%M%S')}.png"
                    output_path = self.captures_dir / filename

                    await self.capture_and_stitch(output_path=output_path, show_success_message=False)

                    current_logs = self.keyword_matches.get(keyword_text, [])
                    status = self._get_status_for_keyword(current_logs)

                    self.report_data.append({
                        'keyword': keyword_text,
                        'lang': keyword_lang,
                        'url': url_str,
                        'status': status or 'N/A',
                        'screenshot_path': output_path
                    })

                # 6. Close browser
                self.update_status(f"Finished with {url_str}. Closing browser.")
                await context.close()
                self.browser_context = None
                self.playwright_page = None

        finally:
            if os.path.exists(user_data_dir):
                try:
                    shutil.rmtree(user_data_dir)
                except Exception as e:
                    print(f"Failed to clean up temp dir: {e}")

    def _select_keyword_programmatically(self, keyword_to_select, event_to_set):
        """Selects a keyword and forces the log view to filter. Must be called from main thread."""
        try:
            all_keyword_texts = self._get_raw_keywords()
            if keyword_to_select in all_keyword_texts:
                idx = all_keyword_texts.index(keyword_to_select)
                
                # Directly set the active filter and update the view
                self.active_filter_keyword = keyword_to_select
                self.keyword_listbox.selection_clear(0, tk.END)
                self.keyword_listbox.selection_set(idx)
                self._refresh_log_view() # This now shows only the filtered logs
        finally:
            event_to_set.set()


    async def wait_for_network_idle(self, idle_duration=3):
        """Waits until no new logs have been added for `idle_duration` seconds."""
        self.update_status("Waiting for network to become idle...")
        while True:
            last_log_count = len(self.all_logs)
            await asyncio.sleep(idle_duration)
            if len(self.all_logs) == last_log_count:
                self.update_status("Network is idle. Proceeding...")
                break

    def start_test_thread(self):
        if not self.playwright_page or self.playwright_page.is_closed():
            messagebox.showwarning("Browser Not Ready", "Please start the browser first.")
            return

        self.test_button.config(state=tk.DISABLED)
        thread = threading.Thread(target=self.run_element_test, daemon=True)
        thread.start()

    def run_element_test(self):
        try:
            future = asyncio.run_coroutine_threadsafe(self.async_run_test(), self.playwright_loop)
            future.result() # Wait for the async test to complete
        except Exception as e:
            print(f"Error during element test: {e}")
        finally:
            self.root.after(0, lambda: self.test_button.config(state=tk.NORMAL))
    
    async def async_run_test(self):
        page = self.playwright_page
        original_url = page.url # Save the main page URL

        modifier = "Meta" if sys.platform == "darwin" else "Control"

        elements_to_test = await page.locator('button, a[href]').all()

        for i, element in enumerate(elements_to_test):
            try:
                if await element.is_visible() and await element.is_enabled():
                    href = await element.get_attribute('href')
                    target = await element.get_attribute('target')

                    is_navigation_link = href and not href.startswith(('#', 'javascript:'))

                    if is_navigation_link and target == '_blank':
                        # Case 1: Link opens a new tab explicitly (target="_blank")
                        await element.click(timeout=5000)
                        await page.wait_for_timeout(2990)
                        await self._close_extra_tabs(original_url)
                        await page.wait_for_timeout(10)

                    elif is_navigation_link:
                        # Case 2: Link navigates in the same tab
                        await element.click(modifiers=[modifier])
                        await page.wait_for_timeout(2990)
                        await self._close_extra_tabs(original_url)
                        await page.wait_for_timeout(10)
                    else:
                        # Case 3: Regular button or javascript link
                        await element.click(timeout=5000)
                        await page.wait_for_timeout(3000)

            except Exception as e:
                print(f"Error clicking element {i+1}: {e}")
        
        # Final cleanup at the end of the test
        await self._close_extra_tabs(original_url)

    async def _close_extra_tabs(self, original_url):
        if not self.browser_context:
            return
        
        pages_to_close = [p for p in self.browser_context.pages if p.url != original_url and not p.is_closed()]
        
        if pages_to_close:
            for p in pages_to_close:
                await p.close()

    async def click_button_by_id(self, button_id):
        if not self.playwright_page or self.playwright_page.is_closed():
            return

        try:
            button_selector = f"#{button_id}"
            element = self.playwright_page.locator(button_selector).first
            
            if await element.count() == 0:
                return

            if not await element.is_visible():
                return

            if not await element.is_enabled():
                return

            await element.click(timeout=5000)

        except Exception as e:
            print(f"Error clicking element with ID '{button_id}': {e}")

    def clear_all(self):
        """Clears all logs and keywords."""
        self.all_logs = []
        self.keyword_listbox.delete(0, tk.END)
        self.keyword_matches = {}
        self.active_filter_keyword = None
        self._refresh_log_view()
        self._save_keyword_state()

    def _on_keyword_click(self, event):
        """Handles clicking on a keyword to load it for editing and toggle filtering."""
        self.is_updating_ui = True
        selection_indices = self.keyword_listbox.curselection()
        
        if not selection_indices:
            # Clicked on empty space: clear inputs and filter
            self.keyword_text_var.set("")
            self.lang_var.set("Traditional Chinese")
            if self.active_filter_keyword:
                self.active_filter_keyword = None
                self._refresh_log_view()
        else:
            # Clicked on a keyword: load it for editing and set filter
            clicked_index = selection_indices[0]
            display_string = self.keyword_listbox.get(clicked_index)
            keyword_obj = self._parse_keyword_display_string(display_string)
            
            self.keyword_text_var.set(keyword_obj['text'])
            self.lang_var.set(self.LANG_MAP_INV.get(keyword_obj['lang'], "Traditional Chinese"))
            self.keyword_num_var.set(str(keyword_obj.get('num', 1)))
            self.button_id_var.set(keyword_obj.get('button_id', ''))

            # Toggle filter logic
            if self.active_filter_keyword == keyword_obj['text']:
                self.active_filter_keyword = None
                self.keyword_listbox.selection_clear(0, tk.END) # Visually deselect
            else:
                self.active_filter_keyword = keyword_obj['text']
            
            self._refresh_log_view()

        self.is_updating_ui = False


    def _refresh_log_view(self):
        """Refreshes the main log tree based on the active filter."""
        for item in self.log_tree.get_children():
            self.log_tree.delete(item)
        
        logs_to_display = []
        if self.active_filter_keyword:
            logs_to_display = self.keyword_matches.get(self.active_filter_keyword, [])
        else:
            logs_to_display = self.all_logs
        
        for log in logs_to_display:
            self.log_tree.insert("", tk.END, values=log)
        
        if not self.active_filter_keyword:
            self.log_tree.yview_moveto(1)

    def sort_treeview(self, col, reverse):
        """Sorts the treeview columns."""
        try:
            data = [(self.log_tree.set(item, col), item) for item in self.log_tree.get_children('')]
            # Attempt to sort numerically if possible
            try:
                data.sort(key=lambda t: float(t[0]), reverse=reverse)
            except ValueError:
                data.sort(key=lambda t: t[0], reverse=reverse)

            for index, (val, item) in enumerate(data):
                self.log_tree.move(item, '', index)

            self.log_tree.heading(col, command=lambda: self.sort_treeview(col, not reverse))
        except Exception as e:
            print(f"Error sorting treeview: {e}")

    def toggle_browser(self):
        if self.browser_button['text'] == "Start Browser":
            self.start_browser_thread()
        else:
            self.close_browser()

    def close_browser(self):
        if self.playwright_page and not self.playwright_page.is_closed():
             # Thread safe close
             try:
                self.playwright_loop.call_soon_threadsafe(lambda: asyncio.create_task(self.playwright_page.close()))
             except Exception as e:
                 print(f"Error closing browser: {e}")

    def start_browser_thread(self):
        display_url = self.url_var.get().strip()
        if not display_url:
            messagebox.showwarning("Input Error", "Please enter or select a URL.")
            return

        # Find the actual URL from the display string
        url_to_run = ""
        try:
            # This handles both direct input and combobox selection
            url_to_run = self._parse_url_display_string(display_url)['url']
        except:
             # If parsing fails, assume the user typed a raw URL
            url_to_run = display_url

        if not url_to_run:
             messagebox.showwarning("Input Error", "Could not determine a valid URL to start.")
             return

        mode = self.mode_var.get()
        
        self.browser_button.config(text="Close Browser")
        
        thread = threading.Thread(target=self.run_playwright, args=(url_to_run, mode), daemon=True)
        thread.start()

    def run_playwright(self, url, mode):
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        self.playwright_loop = loop
        try:
            loop.run_until_complete(self.async_playwright_main(url, mode))
        except Exception as e:
            print(f"Playwright Error: {e}")
            self.root.after(0, lambda: messagebox.showerror("Error", f"Browser Error: {e}"))
        finally:
            self.root.after(0, self.reset_button)
            loop.close()

    def reset_button(self):
        self.browser_button.config(text="Start Browser", state=tk.NORMAL)
        self.playwright_page = None

    async def async_playwright_main(self, url, mode):
        # Create a temporary directory for user data
        user_data_dir = tempfile.mkdtemp()
        
        async with async_playwright() as p:
            # Determine executable path
            chrome_path = "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"
            launch_options = {
                "headless": False,
                "args": [],
            }
            
            if os.path.exists(chrome_path):
                launch_options["executable_path"] = chrome_path
            else:
                print("Custom Chrome path not found, using default Chromium.")

            if mode == "Incognito":
                launch_options["args"].append("--incognito")

            # Launch Persistent Context
            # This is key: launch_persistent_context behaves more like a real user launch
            context = await p.chromium.launch_persistent_context(user_data_dir, **launch_options)
            self.browser_context = context
            
            # Get the default page or create new if none
            page = context.pages[0] if context.pages else await context.new_page()
            self.playwright_page = page

            # Setup Network Interception
            page.on("response", lambda response: self.handle_response(response))

            try:
                print(f"Navigating to {url}")
                await page.goto(url)
                
                # Keep the browser open until closed by user
                # We monitor the close event
                close_event = asyncio.Event()
                page.on("close", lambda: close_event.set())
                
                # Wait until the page is closed
                await close_event.wait()
                
            except Exception as e:
                print(f"Navigation/Runtime Error: {e}")
            finally:
                await context.close()
                # Clean up temporary user data dir
                try:
                    shutil.rmtree(user_data_dir)
                except Exception as cleanup_error:
                    print(f"Failed to clean up temp dir: {cleanup_error}")

    def handle_response(self, response):
        try:
            # Extract relevant data
            url = response.url
            name = url.split('/')[-1]
            if not name:
                name = url

            status = response.status
            method = response.request.method
            resource_type = response.request.resource_type
            size = response.headers.get('content-length', 'N/A')
            timestamp = datetime.now().strftime("%H:%M:%S %d/%m/%Y")

            # Get current URL from the main app and hash it
            current_url = self.url_var.get()
            url_hash = str(hash(current_url))[-8:] if current_url else "N/A"

            # Schedule GUI update on main thread
            # Columns: "name", "status", "method", "type", "size", "time", "url_hash"
            log_values = (name, status, method, resource_type, size, timestamp, url_hash)
            self.root.after(0, self.insert_log, log_values)

        except Exception as e:
            print(f"Error handling response: {e}")

    def insert_log(self, values):
        self.all_logs.append(values)
        
        # If no filter is active, or if the new log matches the active filter, add it to the view
        if not self.active_filter_keyword or self.active_filter_keyword in values[0]:
            self.log_tree.insert("", tk.END, values=values)
            self.log_tree.yview_moveto(1)

        # Debounce the keyword match analysis
        if self.update_timer:
            self.root.after_cancel(self.update_timer)
        self.update_timer = self.root.after(500, self._perform_matching_and_update_list)

    def handle_paste_event(self, event=None):
        """Handles the Ctrl+V/Cmd+V event."""
        self.bulk_add_from_clipboard()

    def handle_keyword_update(self, *args):
        """Automatically updates the selected keyword when its text or lang is changed."""
        if self.is_updating_ui:
            return

        selection_indices = self.keyword_listbox.curselection()
        if not selection_indices:
            return # Nothing selected, nothing to update

        selected_index = selection_indices[0]
        new_text = self.keyword_text_var.get().strip()
        new_lang_display = self.lang_var.get()
        new_lang_short = self.LANG_MAP.get(new_lang_display, "tc")
        new_num = self.keyword_num_var.get().strip()
        new_button_id = self.button_id_var.get().strip()

        if not new_text:
            messagebox.showwarning("Invalid Text", "Keyword text cannot be empty.")
            # Revert to original text
            self.is_updating_ui = True
            original_obj = self._parse_keyword_display_string(self.keyword_listbox.get(selected_index))
            self.keyword_text_var.set(original_obj['text'])
            self.is_updating_ui = False
            return

        # Check for duplicates (text, lang, num) excluding the item being edited
        all_keyword_objects = self._get_keyword_objects()
        for i, obj in enumerate(all_keyword_objects):
            if i != selected_index and \
               obj['text'] == new_text and \
               obj['lang'] == new_lang_short and \
               obj.get('num', 1) == int(new_num):
                messagebox.showwarning("Duplicate Keyword", f"The keyword with this text, lang, and num already exists.")
                # Revert to original text
                self.is_updating_ui = True
                original_obj = self._parse_keyword_display_string(self.keyword_listbox.get(selected_index))
                self.keyword_text_var.set(original_obj['text'])
                self.is_updating_ui = False
                return
        
        # Update the listbox item
        self.is_updating_ui = True # Prevent re-triggering
        original_display = self.keyword_listbox.get(selected_index)
        original_obj = self._parse_keyword_display_string(original_display)
        status_part = ""
        if " (" in original_display:
            status_part = " (" + original_display.split(' (')[-1]
        
        id_part = f" {{{new_button_id}}}" if new_button_id else ""
        new_display_string = f"[{new_num}] [{new_lang_short}] {new_text}{id_part}{status_part}"
        
        self.keyword_listbox.delete(selected_index)
        self.keyword_listbox.insert(selected_index, new_display_string)
        self.keyword_listbox.selection_set(selected_index)
        self.is_updating_ui = False

        self._perform_matching_and_update_list()
        self._save_keyword_state()

    def add_keyword(self):
        """Adds a single keyword from the entry box with its language attribute."""
        keyword_text = self.keyword_text_var.get().strip()
        keyword_lang_display = self.lang_var.get()
        keyword_lang_short = self.LANG_MAP.get(keyword_lang_display, "tc")
        keyword_num = self.keyword_num_var.get().strip()
        button_id = self.button_id_var.get().strip()
        
        if not keyword_text:
            return

        # Prevent duplicates based on text, lang, and num
        new_keyword_obj = {'text': keyword_text, 'lang': keyword_lang_short, 'num': int(keyword_num)}
        all_keyword_objects = self._get_keyword_objects()
        for obj in all_keyword_objects:
            if obj['text'] == new_keyword_obj['text'] and obj['lang'] == new_keyword_obj['lang'] and obj.get('num', 1) == new_keyword_obj['num']:
                messagebox.showwarning("Duplicate", "This exact keyword (text, lang, num) already exists.")
                return

        id_part = f" {{{button_id}}}" if button_id else ""
        display_string = f"[{keyword_num}] [{keyword_lang_short}] {keyword_text}{id_part}"
        self.keyword_listbox.insert(tk.END, display_string)
        
        # Clear inputs for next entry
        self.is_updating_ui = True
        self.keyword_text_var.set("")
        self.is_updating_ui = False
        
        self._perform_matching_and_update_list()
        self._save_keyword_state()

    def bulk_add_from_clipboard(self):
        """Adds multiple keywords from clipboard, using the selected language."""
        try:
            clipboard_content = self.root.clipboard_get()
        except tk.TclError:
            messagebox.showwarning("Paste Error", "Could not read text from clipboard.")
            return
        
        keywords_to_add = clipboard_content.splitlines()
        existing_keywords_text = self._get_raw_keywords()
        selected_lang = self.lang_var.get()
        
        added_count = 0
        for keyword_text in keywords_to_add:
            keyword_text = keyword_text.strip()
            if keyword_text and keyword_text not in existing_keywords_text:
                display_string = f"[{selected_lang}] {keyword_text}"
                self.keyword_listbox.insert(tk.END, display_string)
                existing_keywords_text.append(keyword_text) # Prevent re-adding from same paste
                added_count += 1
        
        if added_count > 0:
            self._perform_matching_and_update_list()
            self._save_keyword_state()

    def remove_selected_keyword_event(self, event=None):
        """Wrapper for keyboard-based deletion of selected keyword."""
        self.remove_selected_keyword()

    def remove_selected_keyword(self):
        """Removes the selected keyword(s) and updates the analysis."""
        selected_indices = self.keyword_listbox.curselection()
        if not selected_indices:
            return
        
        for i in sorted(selected_indices, reverse=True):
            self.keyword_listbox.delete(i)
        
        self.active_filter_keyword = None
        self._perform_matching_and_update_list()
        self._refresh_log_view()
        self._save_keyword_state()

    def remove_all_keywords(self, confirmed=False):
        """Removes all keywords from the list after confirmation."""
        if not self.keyword_listbox.get(0, tk.END):
            return
        
        if confirmed or messagebox.askyesno("Confirm", "Are you sure you want to remove all keywords?"):
            self.keyword_listbox.delete(0, tk.END)
            self.active_filter_keyword = None
            self._perform_matching_and_update_list()
            self._refresh_log_view()
            self._save_keyword_state()

    def _parse_keyword_display_string(self, display_string):
        """Parses '[num] [lang] text {button_id} (status)' into a dictionary."""
        try:
            # New format: [num] [lang] text {button_id} (status)
            num_part, rest = display_string.split('] [', 1)
            num = int(num_part[1:])
            
            lang_part, text_part = rest.split('] ', 1)
            lang = lang_part
            
            # Extract button_id if present
            button_id = ""
            if ' {' in text_part and '}' in text_part:
                text, _, remainder = text_part.partition(' {')
                button_id, _, _ = remainder.partition('}')
            else:
                text = text_part.split(' (')[0].strip()

            return {'text': text.strip(), 'lang': lang.strip(), 'num': num, 'button_id': button_id}
        except (ValueError, IndexError):
            return self._parse_legacy_keyword_string(display_string)

    def _parse_legacy_keyword_string(self, display_string):
        """Fallback parser for old format '[lang] text (status)' or just 'text'."""
        try:
            # Try parsing '[lang] text' format
            parts = display_string.split('] ')
            lang = parts[0][1:]
            text = parts[1].split(' (')[0]
            return {'text': text.strip(), 'lang': lang.strip(), 'num': 1, 'button_id': ''} # Default num and button_id
        except IndexError:
            # If that fails, assume it's just the keyword text
            return {'text': display_string.split(' (')[0].strip(), 'lang': 'tc', 'num': 1, 'button_id': ''}

    def _get_keyword_objects(self):
        """Gets a list of all keyword objects from the listbox."""
        return [self._parse_keyword_display_string(self.keyword_listbox.get(i)) for i in range(self.keyword_listbox.size())]

    def _get_raw_keywords(self):
        """Gets just the text of all keywords."""
        return [obj['text'] for obj in self._get_keyword_objects()]

    def _parse_url_display_string(self, display_string):
        try:
            num_part, rest = display_string.split('] [', 1)
            num = int(num_part[1:])
            lang_part, url_part = rest.split('] ', 1)
            lang = lang_part
            url = url_part
            return {'url': url.strip(), 'lang': lang.strip(), 'num': num}
        except (ValueError, IndexError):
            return {'url': display_string, 'lang': 'tc', 'num': 1}

    def _save_keyword_state(self):
        """Saves the current keyword object list to the undo stack."""
        current_state = self._get_keyword_objects()
        if not self.undo_stack or self.undo_stack[-1] != current_state:
            self.undo_stack.append(current_state)
            self.redo_stack.clear()

    def _restore_keyword_state(self, state):
        """Restores the keyword listbox from a given state of keyword objects."""
        self.keyword_listbox.delete(0, tk.END)
        for keyword_obj in state:
            id_part = f" {{{keyword_obj['button_id']}}}" if keyword_obj.get('button_id') else ""
            display_string = f"[{keyword_obj.get('num', 1)}] [{keyword_obj['lang']}] {keyword_obj['text']}{id_part}"
            self.keyword_listbox.insert(tk.END, display_string)
        self._perform_matching_and_update_list()
        self._refresh_log_view()

    def undo_keywords(self, event=None):
        """Undoes the last keyword change."""
        if len(self.undo_stack) <= 1: # Can't undo past the initial state
            return

        state_for_redo = self.undo_stack.pop()
        self.redo_stack.append(state_for_redo)

        state_to_restore = self.undo_stack[-1]
        self._restore_keyword_state(state_to_restore)

    def redo_keywords(self, event=None):
        """Redoes the last undone keyword change."""
        if not self.redo_stack:
            return

        state_to_restore = self.redo_stack.pop()
        self.undo_stack.append(state_to_restore)
        self._restore_keyword_state(state_to_restore)

    def _get_status_for_keyword(self, logs):
        """Determines the status (PASS, FAILED) for a given list of logs."""
        if not logs:
            return "STANDBY"

        has_failed = False
        all_pass = True

        for log in logs:
            try:
                status_code = int(log[1]) # Status is the second item
                if 400 <= status_code < 500:
                    return "FAILED" # Immediate failure
                elif not (200 <= status_code < 400):
                    # It's a number but not a 2xx or 3xx code (e.g. 5xx).
                    all_pass = False
            except (ValueError, IndexError):
                # Status is not a valid integer or log format is unexpected.
                all_pass = False

        if all_pass:
            return "PASS"
        else:
            # Not a FAILED state, but not a clear PASS state either.
            return None

    def _perform_matching_and_update_list(self):
        """Core function to match logs against keywords and update the listbox UI."""
        self.keyword_matches = {}

        # 1. Get the current view and selection index
        top_fraction, _ = self.keyword_listbox.yview()
        selection_indices = self.keyword_listbox.curselection()
        selected_index = selection_indices[0] if selection_indices else -1

        keyword_objects = self._get_keyword_objects()

        for obj in keyword_objects:
            keyword_text = obj['text']
            self.keyword_matches[keyword_text] = [log for log in self.all_logs if keyword_text in str(log[0])]

        self.keyword_listbox.delete(0, tk.END)
        
        for i, obj in enumerate(keyword_objects):
            keyword_text = obj['text']
            keyword_lang = obj['lang']
            keyword_num = obj.get('num', 1)
            button_id = obj.get('button_id', '')
            
            matched_logs = self.keyword_matches.get(keyword_text, [])
            status = self._get_status_for_keyword(matched_logs)

            id_part = f" {{{button_id}}}" if button_id else ""
            display_string = f"[{keyword_num}] [{keyword_lang}] {keyword_text}{id_part}"
            if status in ("PASS", "FAILED"):
                display_string += f" ({status})"

            self.keyword_listbox.insert(tk.END, display_string)

        # Restore selection and view
        if selected_index != -1:
            self.keyword_listbox.selection_set(selected_index)
        
        self.keyword_listbox.yview_moveto(top_fraction)

    def save_session(self):
        """Saves the current URLs and keyword objects to a JSON file."""
        keywords_to_save = self._get_keyword_objects()
        if not keywords_to_save:
            messagebox.showinfo("Info", "There are no keywords to save.")
            return

        session_data = {
            'urls': self.urls,
            'keywords': keywords_to_save
        }

        file_path = filedialog.asksaveasfilename(
            initialdir=self.sessions_dir,
            defaultextension=".json",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
        )

        if not file_path:
            return

        try:
            with open(file_path, 'w') as f:
                json.dump(session_data, f, indent=4)
            messagebox.showinfo("Success", f"Session saved to {os.path.basename(file_path)}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save session: {e}")

    def load_session(self):
        """Loads URLs and keywords from a JSON file, handling both old and new formats."""
        file_path = filedialog.askopenfilename(
            initialdir=self.sessions_dir,
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
        )

        if not file_path:
            return

        try:
            with open(file_path, 'r') as f:
                session_data = json.load(f)
            
            # --- Load URLs (with backward compatibility) ---
            loaded_urls = []
            if 'urls' in session_data and isinstance(session_data['urls'], list):
                for item in session_data['urls']:
                    if isinstance(item, dict) and 'url' in item and 'lang' in item:
                        if 'num' not in item:
                            item['num'] = 1 # Backward compatibility
                        loaded_urls.append(item)
                    elif isinstance(item, str):
                        loaded_urls.append({'url': item, 'lang': 'tc', 'num': 1}) # Old format
            elif 'url' in session_data: # Even older format
                loaded_urls.append({'url': session_data['url'], 'lang': 'tc', 'num': 1})
            self.update_urls(loaded_urls)
            
            # --- Load Keywords (with backward compatibility) ---
            if 'keywords' in session_data and isinstance(session_data['keywords'], list):
                self.keyword_listbox.delete(0, tk.END)
                for item in session_data['keywords']:
                    if isinstance(item, dict) and 'text' in item and 'lang' in item:
                        item.setdefault('num', 1) # Backward compatibility
                        item.setdefault('button_id', '') # Backward compatibility
                        id_part = f" {{{item['button_id']}}}" if item['button_id'] else ""
                        display_string = f"[{item['num']}] [{item['lang']}] {item['text']}{id_part}"
                        self.keyword_listbox.insert(tk.END, display_string)
                    elif isinstance(item, str):
                        display_string = f"[1] [tc] {item.strip()}"
                        self.keyword_listbox.insert(tk.END, display_string)
                
                self._perform_matching_and_update_list()
                self._save_keyword_state()
                self.update_status(f"Session loaded from {os.path.basename(file_path)}")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to load session: {e}")


class URLManager(tk.Toplevel):
    def __init__(self, parent, app, urls, callback): # Added app parameter
        super().__init__(parent)
        self.app = app # parent is the main app instance
        self.transient(parent)
        self.title("URL Manager")
        self.geometry("700x450")
        self.callback = callback

        self.is_updating_ui = False
        self.undo_stack = deque(maxlen=10)
        self.redo_stack = deque()

        main_frame = ttk.Frame(self, padding=10)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # --- Input Frame ---
        input_frame = ttk.Frame(main_frame)
        input_frame.pack(fill=tk.X, pady=5)

        top_row = ttk.Frame(input_frame)
        top_row.pack(fill=tk.X)
        ttk.Label(top_row, text="URL:").pack(side=tk.LEFT, padx=(0, 5))
        self.url_text_var = tk.StringVar()
        self.url_entry = ttk.Entry(top_row, textvariable=self.url_text_var)
        self.url_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)

        self.url_num_var = tk.StringVar(value="1")
        ttk.Label(top_row, text="Num:").pack(side=tk.LEFT, padx=(5,0))
        self.num_entry = ttk.Entry(top_row, textvariable=self.url_num_var, width=5)
        self.num_entry.pack(side=tk.LEFT, padx=(0,5))

        self.url_lang_var = tk.StringVar(value="Traditional Chinese")
        lang_options = list(self.app.LANG_MAP.keys())
        self.lang_combobox = ttk.Combobox(top_row, textvariable=self.url_lang_var, values=lang_options, state="readonly", width=20)
        self.lang_combobox.pack(side=tk.LEFT, padx=5)

        add_button = ttk.Button(input_frame, text="Add URL", command=self.add_url)
        add_button.pack(fill=tk.X, expand=True, pady=(5, 0))

        # --- Listbox Frame ---
        list_frame = ttk.Frame(main_frame)
        list_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        self.url_listbox = tk.Listbox(list_frame, selectmode=tk.BROWSE)
        self.url_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.url_listbox.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.url_listbox.config(yscrollcommand=scrollbar.set)
        for url_obj in urls:
            self.url_listbox.insert(tk.END, f"[{url_obj.get('num', 1)}] [{url_obj['lang']}] {url_obj['url']}")

        # --- Bindings ---
        self.url_listbox.bind("<ButtonRelease-1>", self._on_url_click)
        self.url_text_var.trace_add("write", self._handle_url_update)
        self.url_lang_var.trace_add("write", self._handle_url_update)
        self.bind("<Control-z>", self.undo_urls)
        self.bind("<Command-z>", self.undo_urls)
        self.bind("<Control-y>", self.redo_urls)
        self.bind("<Command-y>", self.redo_urls)
        self.bind("<Delete>", self.remove_url_event)
        self.bind("<Command-BackSpace>", self.remove_url_event)
        self.bind("<Control-s>", self.save_and_close_event)
        self.bind("<Command-s>", self.save_and_close_event)
        self.bind("<Control-c>", self.copy_selected_url_event)
        self.bind("<Command-c>", self.copy_selected_url_event)
        self.bind("<Control-v>", self.paste_urls_event)
        self.bind("<Command-v>", self.paste_urls_event)

        # --- Button Frame ---
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X)
        remove_button = ttk.Button(button_frame, text="Remove Selected", command=self.remove_url)
        remove_button.pack(side=tk.LEFT)
        save_close_button = ttk.Button(button_frame, text="Save & Close", command=self.save_and_close)
        save_close_button.pack(side=tk.RIGHT)

        self._save_url_state() # Initial state for undo
        self.grab_set()
        self.protocol("WM_DELETE_WINDOW", self.cancel)
        self.wait_window(self)

    def _parse_url_string(self, display_string):
        try:
            num_part, rest = display_string.split('] [', 1)
            num = int(num_part[1:])
            lang_part, url_part = rest.split('] ', 1)
            lang = lang_part
            url = url_part
            return {'url': url.strip(), 'lang': lang.strip(), 'num': num}
        except (ValueError, IndexError):
            # Fallback for old format or direct input
            return {'url': display_string, 'lang': 'tc', 'num': 1}

    def _get_url_objects(self):
        return [self._parse_url_string(self.url_listbox.get(i)) for i in range(self.url_listbox.size())]

    def _on_url_click(self, event):
        self.is_updating_ui = True
        selection_indices = self.url_listbox.curselection()
        if selection_indices:
            display_string = self.url_listbox.get(selection_indices[0])
            url_obj = self._parse_url_string(display_string)
            self.url_text_var.set(url_obj['url'])
            self.url_lang_var.set(self.app.LANG_MAP_INV.get(url_obj['lang'], "Traditional Chinese"))
            self.url_num_var.set(str(url_obj.get('num', 1)))
        self.is_updating_ui = False

    def _handle_url_update(self, *args):
        if self.is_updating_ui:
            return
        selection_indices = self.url_listbox.curselection()
        if not selection_indices:
            return

        selected_index = selection_indices[0]
        new_url = self.url_text_var.get().strip()
        new_lang = self.app.LANG_MAP.get(self.url_lang_var.get(), "tc")
        new_num = self.url_num_var.get().strip()

        if not new_url:
            return # Don't update if URL is empty

        new_display_string = f"[{new_num}] [{new_lang}] {new_url}"
        self.url_listbox.delete(selected_index)
        self.url_listbox.insert(selected_index, new_display_string)
        self.url_listbox.selection_set(selected_index)
        self._save_url_state()

    def add_url(self):
        url = self.url_text_var.get().strip()
        lang = self.app.LANG_MAP.get(self.url_lang_var.get(), "tc")
        num = self.url_num_var.get().strip()
        num = self.url_num_var.get().strip()
        if url:
            display_string = f"[{num}] [{lang}] {url}"
            self.url_listbox.insert(tk.END, display_string)
            self.url_text_var.set("")
            self._save_url_state()

    def remove_url_event(self, event=None):
        self.remove_url()

    def save_and_close_event(self, event=None):
        self.save_and_close()
        return "break" # Prevents the default OS save dialog

    def copy_selected_url_event(self, event=None):
        """Copies the selected URL to the clipboard."""
        selection_indices = self.url_listbox.curselection()
        if not selection_indices:
            return

        selected_url_obj = self._get_url_objects()[selection_indices[0]]
        self.clipboard_clear()
        self.clipboard_append(selected_url_obj['url'])

    def paste_urls_event(self, event=None):
        """Pastes URLs from clipboard into the listbox."""
        try:
            clipboard_content = self.clipboard_get()
        except tk.TclError:
            return # Clipboard is empty or doesn't contain text

        current_urls = {obj['url'] for obj in self._get_url_objects()}
        lang = self.app.LANG_MAP.get(self.url_lang_var.get(), "tc")
        added = False
        for line in clipboard_content.splitlines():
            url = line.strip()
            if url and url not in current_urls:
                display_string = f"[{num}] [{lang}] {url}"
                self.url_listbox.insert(tk.END, display_string)
                current_urls.add(url) # Avoid duplicates within the same paste
                added = True
        
        if added:
            self._save_url_state()

    def remove_url(self):
        selection_indices = self.url_listbox.curselection()
        if selection_indices:
            self.url_listbox.delete(selection_indices[0])
            self._save_url_state()

    def save_and_close(self):
        self.callback(self._get_url_objects())
        self.destroy()

    def cancel(self):
        self.destroy()

    def _save_url_state(self):
        current_state = self._get_url_objects()
        if not self.undo_stack or self.undo_stack[-1] != current_state:
            self.undo_stack.append(current_state)
            self.redo_stack.clear()

    def _restore_url_state(self, state):
        self.url_listbox.delete(0, tk.END)
        for url_obj in state:
            self.url_listbox.insert(tk.END, f"[{url_obj['lang']}] {url_obj['url']}")
        self._save_url_state()

    def undo_urls(self, event=None):
        if len(self.undo_stack) > 1:
            self.redo_stack.append(self.undo_stack.pop())
            self._restore_url_state(self.undo_stack[-1])

    def redo_urls(self, event=None):
        if self.redo_stack:
            state = self.redo_stack.pop()
            self.undo_stack.append(state)
            self._restore_url_state(state)

if __name__ == "__main__":
    root = tk.Tk()
    app = TaggingAutomationApp(root)
    root.mainloop()
